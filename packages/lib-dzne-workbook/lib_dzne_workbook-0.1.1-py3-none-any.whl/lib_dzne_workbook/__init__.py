import openpyxl
import math
import pandas as pd
import sys


def from_DataFrames(dataFrames):
    dataFrames = dict(dataFrames)
    if len(dataFrames) == 0:
        return None
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    for table, df in dataFrames.items():
        if default_sheet is None:
            workbook.create_sheet(table)
        else:
            default_sheet.title = table
            default_sheet = None
    for table, df in dataFrames.items():
        columns = list(df.columns)
        for x, column in enumerate(columns):
            workbook[table].cell(row=1, column=x+1).value = column
            for y, v in enumerate(df[column].tolist()):
                if pd.isna(v):
                    continue
                elif (type(v) is float) and (math.isinf(v)):# is this really needed?
                    value = str(v)
                else:
                    value = v
                workbook[table].cell(row=y+2, column=x+1).value = value
    return workbook

#def write_value(worksheet, value):

